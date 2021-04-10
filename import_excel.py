import sys
from PyQt5 import QtWidgets, QtCore
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QWidget, QMessageBox, QInputDialog
from openpyxl import load_workbook
from pymysql import connect
from pymysql.err import *

host = 'localhost'
port = 3306
user = ''
passwd = ''
database = 'mysql_import'
table = 'output'


def get_column_type(p_type, size):
    if p_type == int:
        return "INT"
    elif p_type == str:
        return "VARCHAR(" + str(size + 10) + ")"
    elif p_type == float:
        return "FLOAT"


def setup_db_connection():
    db = connect(host=host, user=user, password=passwd, database=database, port=port)
    return db


class DatabasePopup(QWidget):
    def __init__(self):
        super().__init__()

        self.database_label = QtWidgets.QLabel(self)
        self.database_edit = QtWidgets.QLineEdit(self)

        self.host_label = QtWidgets.QLabel(self)
        self.host_edit = QtWidgets.QLineEdit(self)

        self.port_label = QtWidgets.QLabel(self)
        self.port_edit = QtWidgets.QLineEdit(self)

        self.table_label = QtWidgets.QLabel(self)
        self.table_edit = QtWidgets.QLineEdit(self)

        self.user_label = QtWidgets.QLabel(self)
        self.user_edit = QtWidgets.QLineEdit(self)

        self.password_label = QtWidgets.QLabel(self)
        self.password_edit = QtWidgets.QLineEdit(self)

        self.submit_label = QtWidgets.QLabel(self)

        self.submit_button = QtWidgets.QPushButton(self)

        self.gridLayout = QtWidgets.QGridLayout(self)
        self.setupUi(self)

    def setupUi(self, MainWindow):
        MainWindow.setObjectName("DB Popup")
        MainWindow.resize(349, 253)

        # self.centralwidget.setAutoFillBackground(True)
        # self.centralwidget.setObjectName("centralwidget")

        self.gridLayout.setObjectName("gridLayout")

        self.database_label.setObjectName("database_label")
        self.gridLayout.addWidget(self.database_label, 0, 0, 1, 2)

        self.database_edit.setObjectName("database_edit")
        self.gridLayout.addWidget(self.database_edit, 0, 2, 1, 1)

        self.host_label.setObjectName("host_label")
        self.gridLayout.addWidget(self.host_label, 1, 0, 1, 2)

        self.host_edit.setObjectName("host_edit")
        self.gridLayout.addWidget(self.host_edit, 1, 2, 1, 1)

        self.port_label.setObjectName("port_label")
        self.gridLayout.addWidget(self.port_label, 2, 0, 1, 2)

        self.port_edit.setObjectName("port_edit")
        self.gridLayout.addWidget(self.port_edit, 2, 2, 1, 1)

        self.table_label.setObjectName("table_label")
        self.gridLayout.addWidget(self.table_label, 3, 0, 1, 2)

        self.table_edit.setObjectName("table_edit")
        self.gridLayout.addWidget(self.table_edit, 3, 2, 1, 1)

        self.user_label.setObjectName("user_label")
        self.gridLayout.addWidget(self.user_label, 4, 0, 1, 2)

        self.user_edit.setObjectName("user_edit")
        self.gridLayout.addWidget(self.user_edit, 4, 2, 1, 1)

        self.password_label.setObjectName("password_label")
        self.gridLayout.addWidget(self.password_label, 5, 0, 1, 2)

        self.password_edit.setObjectName("password_edit")
        self.gridLayout.addWidget(self.password_edit, 5, 2, 1, 1)

        self.submit_label.setObjectName("submit_label")
        self.gridLayout.addWidget(self.submit_label, 6, 0, 1, 3)

        self.submit_button.setObjectName("submit_button")
        self.gridLayout.addWidget(self.submit_button, 7, 0, 1, 3)
        self.submit_button.clicked.connect(self.submit_clicked)

        self.database_label.raise_()
        self.database_edit.raise_()
        self.host_label.raise_()
        self.host_edit.raise_()

        self.setLayout(self.gridLayout)
        # self.setCentralWidget(self.centralwidget)

        # MainWindow.setCentralWidget(self.centralwidget)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)
        self.show()

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "DB Credentials"))
        self.database_label.setText(_translate("MainWindow", "Database: "))
        self.database_edit.setText(_translate("MainWindow", database))
        self.host_label.setText(_translate("MainWindow", "Host: "))
        self.host_edit.setText(_translate("MainWindow", host))
        self.port_label.setText(_translate("MainWindow", "Port: "))
        self.port_edit.setText(_translate("MainWindow", str(port)))
        self.table_label.setText(_translate("MainWindow", "Table: "))
        self.table_edit.setText(_translate("MainWindow", table))
        self.user_label.setText(_translate("MainWindow", "User: "))
        self.user_edit.setText(_translate("MainWindow", user))
        self.password_label.setText(_translate("MainWindow", "Password: "))
        self.password_edit.setText(_translate("MainWindow", passwd))
        self.submit_label.setText(_translate("MainWindow", ""))
        self.submit_button.setText(_translate("MainWindow", "Submit"))

    def submit_clicked(self):
        global database
        database = self.database_edit.text()
        global host
        host = self.host_edit.text()
        global port
        port = int(self.port_edit.text())
        global table
        table = self.table_edit.text()
        global user
        user = self.user_edit.text()
        global passwd
        passwd = self.password_edit.text()
        try:
            setup_db_connection()
            self.close()
        except Exception as ex:
            self.submit_label.setText(str(ex))


class DragAndDropPopup(QWidget):
    def __init__(self, ui, labels):
        super().__init__()
        self.parent_ui = ui
        self.columns_label = QtWidgets.QLabel(self)
        self.columns_list = QtWidgets.QListWidget(self)
        self.gridLayout = QtWidgets.QGridLayout(self)
        self.submit_buttom = QtWidgets.QPushButton(self)
        self.setupUi(self, labels)

    def setupUi(self, MainWindow, labels):
        MainWindow.setObjectName("Drag and Drop Popup")
        MainWindow.resize(349, 253)

        self.columns_list.setObjectName("columns_list")
        for key in range(len(labels)):
            self.columns_list.insertItem(key, labels[key])
        self.columns_list.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        self.gridLayout.addWidget(self.columns_list, 1, 0, 1, 3)

        # self.columns_edit.setText(labels)
        # self.columns_edit.setReadOnly(True)
        # self.columns_edit.setDragEnabled(True)
        # self.columns_edit.setObjectName("column_values")
        # self.gridLayout.addWidget(self.columns_edit, 1, 0, 1, 3)

        # self.import_columns_edit.setText("")
        # self.import_columns_edit.setDragEnabled(False)
        # self.import_columns_edit.setObjectName("import_columns_edit")
        # self.gridLayout.addWidget(self.import_columns_edit, 2, 0, 1, 3)

        self.submit_buttom.setObjectName("custom_import_button")
        self.gridLayout.addWidget(self.submit_buttom, 3, 0, 1, 3)
        self.submit_buttom.clicked.connect(self.get_columns)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)
        self.show()

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Custom Columns"))
        self.submit_buttom.setText(_translate("MainWindow", "Submit"))

    def get_columns(self):
        selected_columns = ""
        for key in self.columns_list.selectedItems():
            selected_columns += key.text() + ","
        self.close()
        self.parent_ui.import_custom_values(selected_columns)


class Ui(QMainWindow):
    def __init__(self):
        super(Ui, self).__init__()
        self.headings = []
        self.db_connection = None
        self.data = []
        self.create_table = ""
        self.centralwidget = QtWidgets.QWidget(self)
        self.upload_edit = QtWidgets.QLineEdit(self.centralwidget)
        self.upload_button = QtWidgets.QPushButton(self.centralwidget)
        self.all_columns_label = QtWidgets.QLabel(self.centralwidget)

        self.text_edit = QtWidgets.QTextEdit(self.centralwidget)

        self.custom_import_button = QtWidgets.QPushButton(self.centralwidget)
        self.all_import_button = QtWidgets.QPushButton(self.centralwidget)

        self.gridLayout = QtWidgets.QGridLayout(self.centralwidget)
        self.setupUi(self)

        self.show()  # Show the GUI

    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(349, 253)

        self.centralwidget.setAutoFillBackground(True)
        self.centralwidget.setObjectName("centralwidget")

        self.gridLayout.setObjectName("gridLayout")

        self.upload_edit.setReadOnly(True)
        self.upload_edit.setObjectName("upload_edit")
        self.gridLayout.addWidget(self.upload_edit, 0, 0, 1, 2)

        self.upload_button.setObjectName("upload_button")
        self.upload_button.clicked.connect(self.upload_excel_data)
        self.gridLayout.addWidget(self.upload_button, 0, 2, 1, 1)

        self.all_columns_label.setText("Your column description:")
        self.all_columns_label.setObjectName("all_columns_label")
        self.gridLayout.addWidget(self.all_columns_label, 1, 0, 1, 2)

        self.text_edit.setText("")
        self.text_edit.setReadOnly(True)
        self.text_edit.setObjectName("column_values")
        self.gridLayout.addWidget(self.text_edit, 2, 0, 1, 3)

        self.all_import_button.setObjectName("all_import_button")
        self.gridLayout.addWidget(self.all_import_button, 3, 0, 1, 2)
        self.all_import_button.clicked.connect(self.import_all)

        self.custom_import_button.setObjectName("custom_import_button")
        self.gridLayout.addWidget(self.custom_import_button, 3, 2, 1, 1)
        self.custom_import_button.clicked.connect(self.import_custom)

        self.upload_button.raise_()
        self.upload_edit.raise_()
        self.text_edit.raise_()
        self.custom_import_button.raise_()
        self.all_import_button.raise_()
        MainWindow.setCentralWidget(self.centralwidget)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Excel to MySQL"))
        self.upload_button.setText(_translate("MainWindow", "Upload"))
        self.custom_import_button.setText(_translate("MainWindow", "Custom Import"))
        self.all_import_button.setText(_translate("MainWindow", "Import All"))

    def upload_excel_data(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getOpenFileName(self, "QFileDialog.getOpenFileName()", "",
                                                  "Excel Files (*.xlsx);;All Files (*)", options=options)
        if fileName:
            self.upload_edit.setText(fileName.split("/")[-1])
        db_column_size = self.read_excel_headers(self.upload_edit.text())
        self.set_create_table(db_column_size)

    def select_sheet_name(self, sheetnames):
        if len(sheetnames) > 1:
            item, ok = QInputDialog.getItem(self, "select input dialog",
                                            "list of sheets", sheetnames, 0, False)
            if item and ok:
                return item

    def read_excel_data(self, path, sheet_name, db_column_size):

        wb_obj = load_workbook(path)
        if sheet_name:
            sheet_obj = wb_obj[sheet_name]
        else:
            sheet_obj = wb_obj.active
        max_col = sheet_obj.max_column
        max_row = sheet_obj.max_row

        for i in range(1, max_row):
            temp_dict = {}
            for j in range(0, max_col):
                cell_obj = sheet_obj.cell(row=i + 1, column=j + 1)
                temp_dict.update({self.headings[j]: cell_obj.value})
                if type(cell_obj.value) == str:
                    if len(cell_obj.value) > db_column_size[j]:
                        db_column_size[j] = len(cell_obj.value)
            self.data.append(temp_dict)

    def read_excel_headers(self, path):
        # workbook object is created
        wb_obj = load_workbook(path)

        sheet_name = self.select_sheet_name(wb_obj.sheetnames)
        if sheet_name:
            sheet_obj = wb_obj[sheet_name]
        else:
            sheet_obj = wb_obj.active
        max_col = sheet_obj.max_column
        max_row = sheet_obj.max_row

        # Loop will print all columns name
        db_column_size = []

        for i in range(0, max_col):
            cell_obj = sheet_obj.cell(row=1, column=i + 1)
            self.headings.append(cell_obj.value)
            db_column_size.append(0)

        for i in range(1, max_row):
            temp_dict = {}
            for j in range(0, max_col):
                cell_obj = sheet_obj.cell(row=i + 1, column=j + 1)
                temp_dict.update({self.headings[j]: cell_obj.value})
                if type(cell_obj.value) == str:
                    if len(cell_obj.value) > db_column_size[j]:
                        db_column_size[j] = len(cell_obj.value)
            self.data.append(temp_dict)
        self.set_label_headings()
        return db_column_size

    def import_all(self):
        if self.db_connection is None:
            if not database or not host or not port or not table or not user or not passwd:
                self.open_db_dialog()
            else:
                self.import_all_values()
        else:
            self.import_all_values()

    def execute_create_table(self):
        self.db_connection = setup_db_connection()
        cur = self.db_connection.cursor()
        try:
            cur.execute("show create table " + table + ";")
            table_def = cur.fetchone()
        except ProgrammingError:
            cur.execute(self.create_table)
            self.db_connection.commit()
            return
        lowered_headings = []
        for key in self.headings:
            lowered_headings.append(key.lower())

        if table_def:
            if not all(table_def[1].split('`').__contains__(key) for key in lowered_headings):
                cur.execute("drop table if exists " + table + ";")
            else:
                return
        cur.execute(self.create_table)
        self.db_connection.commit()

    def import_all_values(self):
        self.execute_create_table()
        self.db_connection = setup_db_connection()
        cur = self.db_connection.cursor()

        for key in range(len(self.data)):
            query = "INSERT INTO " + table + " VALUES("
            values = ""
            for i in range(len(self.headings)):
                if type(self.data[key][self.headings[i]]) == str:
                    values += "'" + str(self.data[key][self.headings[i]]) + "', "
                else:
                    values += str(self.data[key][self.headings[i]]) + ", "
            query += values[:-2] + ");"
            cur.execute(query)
        self.db_connection.commit()
        self.db_connection.close()
        self.buildPopup("Import All Status", "Sucessfully imported " + str(len(self.data)) + " entries.")

    def open_db_dialog(self):
        self.d = DatabasePopup()

    def import_custom(self):
        if self.db_connection is None:
            if not database or not host or not port or not table or not user or not passwd:
                self.open_db_dialog()
        self.dnd = DragAndDropPopup(self, self.headings)

    def import_custom_values(self, text):
        self.execute_create_table()
        self.db_connection = setup_db_connection()
        cur = self.db_connection.cursor()

        heads = text.strip('\n\t ,').split(",")
        heads_str = ""
        final_heads = []

        for key in range(len(heads)):
            final_heads.append(heads[key].upper().strip('\n\t ,'))
            heads_str += heads[key] + ", "

        for key in range(len(self.data)):
            query = "INSERT INTO " + table + "("

            query += heads_str[:-2]
            values = ") VALUES("
            for i in range(len(self.headings)):
                if final_heads.__contains__(self.headings[i]):
                    if type(self.data[key][self.headings[i]]) == str:
                        values += "'" + str(self.data[key][self.headings[i]]) + "', "
                    else:
                        values += str(self.data[key][self.headings[i]]) + ", "
            query += values[:-2] + ");"
            print(query)
            try:
                cur.execute(query)
            except Exception as ex:
                self.buildPopup("Import Customer Status",
                                str(ex),
                                "exception")
                return
            self.db_connection.commit()
        self.buildPopup("Import Customer Status",
                        "Sucessfully imported " + str(heads) + " " + str(len(self.data)) + " entries.")
        self.db_connection.close()

    def set_create_table(self, column_sizes):
        query_table = "CREATE TABLE " + table + " ("
        for key in range(len(self.headings)):
            query_table += self.headings[key].lower() + " " + get_column_type(type(self.data[0][self.headings[key]]),
                                                                              column_sizes[key]) + ","
        self.create_table = query_table[:-1] + ");"

    def set_label_headings(self):
        heading_list = ""
        for key in self.headings:
            heading_list += key + ",\n"
        self.text_edit.setText(heading_list[:-2])

    def get_label_headings(self):
        heading_list = ""
        for key in self.headings:
            heading_list += key + ",\n"
        return heading_list[:-2]

    def buildPopup(self, name, text, type='information'):
        msg = QMessageBox()
        msg.setWindowTitle(name)
        msg.setText(text)
        if type == 'information':
            msg.setIcon(QMessageBox.Information)
        elif type == 'exception' or type == 'error' or type == 'critical' or type == 'danger':
            msg.setIcon(QMessageBox.Critical)
        x = msg.exec_()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    demo = Ui()
    sys.exit(app.exec_())
